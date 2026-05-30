#!/usr/bin/env python3
"""
DCF demo — follows the dcf-model + xlsx-author skills:
- Inputs tab holds every hardcoded assumption (blue font).
- DCF / WACC / Sensitivity tabs are 100% live formulas (black font).
- No market-data MCP: all assumptions typed by hand.
Run:  python3 build_dcf.py   ->  ./out/dcf-DEMO.xlsx
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

BLUE = Font(color="0000FF")          # hardcoded input
BLACK = Font(color="000000")         # formula
BOLD = Font(bold=True)
PCT = "0.0%"
NUM = "#,##0.0"

wb = Workbook()

# ---------------- Inputs ----------------
inp = wb.active
inp.title = "Inputs"
inp["B1"] = "DEMO Corp — DCF inputs ($M unless noted)"; inp["B1"].font = BOLD
rows = [
    ("B3", "Latest revenue (FY0)", "C3", 1000),
    ("B4", "Diluted shares (M)",   "C4", 100),
    ("B5", "Net debt",             "C5", 200),
    ("B6", "Current price ($)",    "C6", 50),
    ("B7", "Tax rate",             "C7", 0.25),
]
for lb, lt, cc, v in rows:
    inp[lb] = lt
    inp[cc] = v; inp[cc].font = BLUE
inp["C7"].number_format = PCT

inp["B9"] = "Year"; inp["B9"].font = BOLD
years = [1, 2, 3, 4, 5]
cols = ["C", "D", "E", "F", "G"]
for c, y in zip(cols, years):
    inp[f"{c}9"] = y; inp[f"{c}9"].font = BOLD

drivers = [
    (10, "Revenue growth", [0.16, 0.14, 0.12, 0.10, 0.09]),
    (11, "EBIT margin",    [0.25, 0.26, 0.27, 0.28, 0.28]),
    (12, "D&A % rev",      [0.04, 0.04, 0.04, 0.04, 0.04]),
    (13, "Capex % rev",    [0.05, 0.05, 0.05, 0.05, 0.05]),
    (14, "dNWC % of dRev", [0.10, 0.10, 0.10, 0.10, 0.10]),
]
for r, lab, vals in drivers:
    inp[f"B{r}"] = lab
    for c, v in zip(cols, vals):
        inp[f"{c}{r}"] = v; inp[f"{c}{r}"].font = BLUE; inp[f"{c}{r}"].number_format = PCT
inp.column_dimensions["B"].width = 24

# ---------------- DCF ----------------
dcf = wb.create_sheet("DCF")
dcf["B1"] = "DCF — DEMO Corp ($M)"; dcf["B1"].font = BOLD
# discounting assumptions (blue inputs, lives on DCF sheet)
dcf["B3"] = "WACC";            dcf["C3"] = 0.09; dcf["C3"].font = BLUE; dcf["C3"].number_format = PCT
dcf["B4"] = "Terminal growth"; dcf["C4"] = 0.03; dcf["C4"].font = BLUE; dcf["C4"].number_format = PCT

dcf["B6"] = "Year"; dcf["B6"].font = BOLD
for c, y in zip(cols, years):
    dcf[f"{c}6"] = y; dcf[f"{c}6"].font = BOLD

def line(r, label, fml_by_col, fmt=NUM):
    dcf[f"B{r}"] = label
    for i, c in enumerate(cols):
        dcf[f"{c}{r}"] = fml_by_col(i, c); dcf[f"{c}{r}"].font = BLACK
        dcf[f"{c}{r}"].number_format = fmt

# Revenue: FY1 from Inputs FY0; later years compound on prior DCF revenue
def rev(i, c):
    if i == 0:
        return f"=Inputs!C3*(1+Inputs!C10)"
    prev = cols[i-1]
    return f"={prev}7*(1+Inputs!{c}10)"
line(7, "Revenue", rev)
line(8, "EBIT", lambda i, c: f"={c}7*Inputs!{c}11")
line(9, "NOPAT", lambda i, c: f"={c}8*(1-Inputs!$C$7)")
line(10, "D&A", lambda i, c: f"={c}7*Inputs!{c}12")
line(11, "Capex", lambda i, c: f"={c}7*Inputs!{c}13")
def dnwc(i, c):
    prevrev = "Inputs!C3" if i == 0 else f"{cols[i-1]}7"
    return f"=({c}7-{prevrev})*Inputs!{c}14"
line(12, "Change in NWC", dnwc)
line(13, "Unlevered FCF", lambda i, c: f"={c}9+{c}10-{c}11-{c}12")
line(14, "Discount factor", lambda i, c: f"=1/(1+$C$3)^{c}6", fmt="0.000")
line(15, "PV of FCF", lambda i, c: f"={c}13*{c}14")

dcf["B17"] = "Sum PV of FCF";        dcf["C17"] = "=SUM(C15:G15)"; dcf["C17"].number_format = NUM
dcf["B18"] = "Terminal value (FY5)"; dcf["C18"] = "=G13*(1+$C$4)/($C$3-$C$4)"; dcf["C18"].number_format = NUM
dcf["B19"] = "PV of terminal value"; dcf["C19"] = "=C18*G14"; dcf["C19"].number_format = NUM
dcf["B20"] = "Enterprise value";     dcf["C20"] = "=C17+C19"; dcf["C20"].number_format = NUM
dcf["B21"] = "Less: net debt";       dcf["C21"] = "=-Inputs!C5"; dcf["C21"].number_format = NUM
dcf["B22"] = "Equity value";         dcf["C22"] = "=C20+C21"; dcf["C22"].number_format = NUM
dcf["B23"] = "Diluted shares";       dcf["C23"] = "=Inputs!C4"; dcf["C23"].number_format = NUM
dcf["B24"] = "Value per share ($)";  dcf["C24"] = "=C22/C23"; dcf["C24"].number_format = "#,##0.00"
dcf["B25"] = "Current price ($)";    dcf["C25"] = "=Inputs!C6"; dcf["C25"].number_format = "#,##0.00"
dcf["B26"] = "Upside / (downside)";  dcf["C26"] = "=C24/C25-1"; dcf["C26"].number_format = PCT
for r in range(17, 27):
    dcf[f"B{r}"].font = BOLD if r in (20, 24, 26) else BLACK
    dcf[f"C{r}"].font = BLACK
dcf.column_dimensions["B"].width = 22

# ---------------- WACC (bottom-up build) ----------------
w = wb.create_sheet("WACC")
w["B1"] = "WACC build (illustrative)"; w["B1"].font = BOLD
wrows = [
    ("Risk-free rate", 0.042, True, PCT),
    ("Beta", 1.10, True, "0.00"),
    ("Equity risk premium", 0.055, True, PCT),
    ("Cost of equity", "=C3+C4*C5", False, PCT),
    ("Pre-tax cost of debt", 0.055, True, PCT),
    ("Tax rate", 0.25, True, PCT),
    ("After-tax cost of debt", "=C7*(1-C8)", False, PCT),
    ("Equity weight", 0.80, True, PCT),
    ("Debt weight", 0.20, True, PCT),
    ("WACC", "=C6*C9+C10*C11", False, PCT),
]
for i, (lab, val, is_input, fmt) in enumerate(wrows):
    r = i + 3
    w[f"B{r}"] = lab
    w[f"C{r}"] = val
    w[f"C{r}"].font = BLUE if is_input else BLACK
    w[f"C{r}"].number_format = fmt
w["B12"].font = BOLD; w["C12"].font = BOLD
w.column_dimensions["B"].width = 24

# ---------------- Sensitivity (live 5x5) ----------------
s = wb.create_sheet("Sensitivity")
s["B1"] = "Value per share ($) — WACC (rows) x terminal growth (cols)"; s["B1"].font = BOLD
s["B3"] = "WACC \\ g"; s["B3"].font = BOLD
g_headers = [0.02, 0.025, 0.03, 0.035, 0.04]
w_headers = [0.08, 0.085, 0.09, 0.095, 0.10]
gcols = ["C", "D", "E", "F", "G"]
for c, g in zip(gcols, g_headers):
    s[f"{c}3"] = g; s[f"{c}3"].font = BLUE; s[f"{c}3"].number_format = PCT
for ri, wv in enumerate(w_headers):
    r = 4 + ri
    s[f"B{r}"] = wv; s[f"B{r}"].font = BLUE; s[f"B{r}"].number_format = PCT
    for c in gcols:
        # recompute value/share with this cell's WACC ($B{r}) and g ({c}$3),
        # discounting the fixed FCF stream from the DCF sheet (DCF!C13:G13).
        pv_fcf = "+".join(f"DCF!${col}$13/(1+$B{r})^{k}" for k, col in zip(years, cols))
        pv_tv = f"(DCF!$G$13*(1+{c}$3)/($B{r}-{c}$3))/(1+$B{r})^5"
        s[f"{c}{r}"] = f"=(({pv_fcf}+{pv_tv})-Inputs!$C$5)/Inputs!$C$4"
        s[f"{c}{r}"].font = BLACK; s[f"{c}{r}"].number_format = "#,##0.00"
s.column_dimensions["B"].width = 10

os.makedirs("./out", exist_ok=True)
path = "./out/dcf-DEMO.xlsx"
wb.save(path)
print("wrote", path)
print("sheets:", wb.sheetnames)
