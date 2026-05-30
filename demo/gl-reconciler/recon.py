#!/usr/bin/env python3
"""
GL <-> subledger reconciliation — reference implementation of the gl-recon skill
(Step 1 normalize, Step 2 full-outer-join match, Step 3 classify cause, Step 4 report).
This is what the GL Reconciler agent produces via reasoning + xlsx-author, written
out here as a script so the offline demo is reproducible. No MCP, no network.

Run:  python3 recon.py   ->  ./out/break-report.xlsx  (+ console summary)
"""
import csv, os
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

KEY = ("security_id", "account", "trade_date")
TOL_AMT = 0.01   # default amount tolerance from the skill
TOL_QTY = 0.0

def load(path):
    out = OrderedDict()
    with open(path) as f:
        for row in csv.DictReader(f):
            k = tuple(row[c].strip().upper() for c in KEY)          # Step 1: upper-stripped key
            for c in ("quantity", "local_amount", "base_amount", "fx_rate"):
                row[c] = round(float(row[c]), 2)                    # Step 1: numeric coercion
            out[k] = row
    return out

gl, sub = load("gl.csv"), load("subledger.csv")

def classify(g, s):
    """Step 2 bucket + Step 3 likely cause. Returns (bucket, cause, note)."""
    if s is None:
        return "GL only", "Duplicate / missing post", "Present in GL, absent in subledger"
    if g is None:
        return "Subledger only", "Duplicate / missing post", "Present in subledger, absent in GL"
    qty_eq = abs(g["quantity"] - s["quantity"]) <= TOL_QTY
    base_eq = abs(g["base_amount"] - s["base_amount"]) <= TOL_AMT
    local_eq = abs(g["local_amount"] - s["local_amount"]) <= TOL_AMT
    date_eq = g["posting_date"] == s["posting_date"]
    if qty_eq and base_eq and local_eq and date_eq:
        return "Matched", "", ""
    if qty_eq and base_eq and local_eq and not date_eq:
        return "Timing break", "Timing", f"Posting dates differ ({g['posting_date']} vs {s['posting_date']}); amounts agree"
    if not qty_eq:
        return "Quantity break", "Data quality", f"Quantity differs ({g['quantity']} vs {s['quantity']})"
    if local_eq and not base_eq:
        return "Amount break", "FX", f"Local amounts agree, base differ ({g['base_amount']} vs {s['base_amount']}); fx {g['fx_rate']} vs {s['fx_rate']}"
    return "Amount break", "Mapping", "Amount differs; cause to confirm"

rows = []
for k in list(gl.keys()) + [k for k in sub if k not in gl]:      # Step 2: full outer join
    g, s = gl.get(k), sub.get(k)
    bucket, cause, note = classify(g, s)
    if bucket == "Matched":
        rows.append((k, g, s, bucket, cause, note, 0.0)); continue
    gb = g["base_amount"] if g else 0.0
    sb = s["base_amount"] if s else 0.0
    rows.append((k, g, s, bucket, cause, note, abs(gb - sb)))

breaks = [r for r in rows if r[3] != "Matched"]
breaks.sort(key=lambda r: r[6], reverse=True)                    # Step 4: sort by |base delta| desc
matched = [r for r in rows if r[3] == "Matched"]

# ---- console summary (Step 4 summary) ----
print(f"Rows compared: {len(rows)}  |  Matched: {len(matched)}  |  Breaks: {len(breaks)}")
from collections import Counter
print("By bucket:", dict(Counter(r[3] for r in breaks)))
print(f"Matched %: {len(matched)/len(rows):.0%}")
for r in breaks:
    print(f"  {r[0][0]:7} {r[3]:16} cause={r[4]:14} |Δbase|={r[6]:.2f}  {r[5]}")

# ---- Step 4: break report workbook (xlsx-author conventions) ----
wb = Workbook()
ws = wb.active; ws.title = "Break report"
HDR = Font(bold=True, color="FFFFFF"); FILL = PatternFill("solid", fgColor="305496")
head = ["security_id","account","trade_date","bucket","likely_cause",
        "gl_base","sub_base","base_delta","note"]
ws.append(head)
for c in range(1, len(head)+1):
    ws.cell(1, c).font = HDR; ws.cell(1, c).fill = FILL
for k, g, s, bucket, cause, note, delta in breaks:
    ws.append([k[0], k[1], k[2], bucket, cause,
               g["base_amount"] if g else None,
               s["base_amount"] if s else None,
               round(delta, 2), note])
for col, wdt in zip("ABCDEFGHI", (12,9,12,16,16,12,12,12,52)):
    ws.column_dimensions[col].width = wdt

sm = wb.create_sheet("Summary")
sm.append(["Reconciliation summary — trade_date 2026-05-05, equities"])
sm["A1"].font = Font(bold=True)
sm.append(["Rows compared", len(rows)])
sm.append(["Matched", len(matched)])
sm.append(["Breaks", len(breaks)])
sm.append(["Matched %", round(len(matched)/len(rows), 4)])
sm["B6"] = ""
sm.append([]); sm.append(["By bucket", "count"])
for b, n in Counter(r[3] for r in breaks).items():
    sm.append([b, n])
sm.append([]); sm.append(["GUARDRAIL", "Report only — ledger adjustments require human controller sign-off (No ledger posting)."])
sm.column_dimensions["A"].width = 22; sm.column_dimensions["B"].width = 60

os.makedirs("./out", exist_ok=True)
wb.save("./out/break-report.xlsx")
print("wrote ./out/break-report.xlsx")
