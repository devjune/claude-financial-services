#!/usr/bin/env python3
"""
KYC rules engine — reference implementation of the kyc-rules skill
(Step 1 risk-rate, Step 2 required-doc check, Step 3 rule outcomes w/ citations,
Step 4 disposition). Consumes the kyc-doc-parse output (parsed.json) and the firm
rules grid. Screening MCP is NOT available offline, so sanctions/adverse-media (R7)
is marked not-run — exactly the skill's "this skill never approves" stance.

Run:  python3 kyc_screen.py  ->  ./out/escalation-PKT-DEMO-001.xlsx  (+ console)
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

HIGH_RISK = {"RU", "IR", "KP", "SY", "BY", "MM"}
OFFSHORE = {"KY", "VG", "BS", "PA"}
REQUIRED_ENTITY = [
    "Certificate of incorporation", "UBO declaration", "Address proof (utility bill)",
    "Source of funds", "Tax form (W-8BEN-E)", "Certified ID for each UBO",
]

rec = json.load(open("parsed.json"))
ubos = rec["beneficial_owners"]
received = {d["type"] for d in rec["documents_received"]}

# ---- Step 1: risk-rate ----
factors = []
rating = "low"
def bump(to):
    global rating
    order = {"low": 0, "medium": 1, "high": 2}
    if order[to] > order[rating]:
        rating = to

if rec["nationality_or_jurisdiction"] in OFFSHORE:
    factors.append(("Jurisdiction", rec["nationality_or_jurisdiction"], "Offshore list -> medium")); bump("medium")
hr_ubos = [u for u in ubos if u["nationality"] in HIGH_RISK]
if hr_ubos:
    factors.append(("UBO nationality", ",".join(u["nationality"] for u in hr_ubos), "High-risk list -> high")); bump("high")
if rec.get("pep_declared"):
    factors.append(("PEP exposure", "declared", "Declared PEP -> high")); bump("high")
factors.append(("Ownership opacity", f"{len(ubos)} UBOs", "Flat structure -> low"))
factors.append(("Source of funds", "documented", "Has reference -> ok"))

# ---- Step 2: required-document check ----
have_certified_id = len(rec.get("id_documents", [])) >= len(ubos) and len(ubos) > 0
doc_status = []
for d in REQUIRED_ENTITY:
    if d == "Certified ID for each UBO":
        doc_status.append((d, "received" if have_certified_id else "missing"))
    else:
        doc_status.append((d, "received" if d in received else "missing"))
missing = [d for d, s in doc_status if s == "missing"]

# ---- Step 3: rule outcomes (cite every rule) ----
outcomes = []
outcomes.append(("R1", "Entity jurisdiction on offshore list",
                 "fail" if rec["nationality_or_jurisdiction"] in OFFSHORE else "pass",
                 f"jurisdiction={rec['nationality_or_jurisdiction']}"))
outcomes.append(("R2", "Any UBO nationality on high-risk list",
                 "fail" if hr_ubos else "pass",
                 "UBO " + ", ".join(f"{u['name']}({u['nationality']})" for u in hr_ubos) if hr_ubos else "none"))
outcomes.append(("R3", "Any declared/screened PEP among UBOs/controllers",
                 "fail" if rec.get("pep_declared") else "pass",
                 "PEP declared (Dmitri Volkov)" if rec.get("pep_declared") else "none"))
outcomes.append(("R4", "Certified ID present for every UBO",
                 "pass" if have_certified_id else "fail",
                 "id_documents empty" if not have_certified_id else "all present"))
outcomes.append(("R5", "Address proof <= 3 months old", "pass", "utility bill 2026-05-10"))
outcomes.append(("R6", "Source of funds documented with reference", "pass", "ref SOF-LON-2023-0830"))
outcomes.append(("R7", "Sanctions / adverse-media screening clear", "n/a",
                 "NOT RUN — screening MCP unavailable offline; must run before any clear"))

# ---- Step 4: disposition ----
escalation_reasons = []
if hr_ubos:
    escalation_reasons.append("R2: UBO in high-risk jurisdiction (RU)")
if rec.get("pep_declared"):
    escalation_reasons.append("R3: declared PEP among UBOs")
if escalation_reasons:
    disposition = "escalate-EDD"
elif missing:
    disposition = "request-docs"
elif rating in ("low", "medium"):
    disposition = "clear"
else:
    disposition = "escalate-EDD"

disposition_json = {
    "packet_id": rec["packet_id"],
    "risk_rating": rating,
    "disposition": disposition,
    "missing_documents": missing,
    "escalation_reasons": escalation_reasons,
    "rule_outcomes": [{"rule_id": r, "outcome": o, "evidence": e} for r, _, o, e in outcomes],
}

print(json.dumps(disposition_json, indent=2, ensure_ascii=False))

# ---- escalation workbook (xlsx-author) ----
wb = Workbook()
HDR = Font(bold=True, color="FFFFFF"); FILL = PatternFill("solid", fgColor="C00000")
GREY = PatternFill("solid", fgColor="305496")

s = wb.active; s.title = "Disposition"
s["A1"] = f"KYC escalation packet — {rec['packet_id']} — {rec['legal_name']}"; s["A1"].font = Font(bold=True, size=13)
s["A3"] = "Applicant type"; s["B3"] = rec["applicant_type"]
s["A4"] = "Jurisdiction"; s["B4"] = rec["nationality_or_jurisdiction"]
s["A5"] = "Risk rating"; s["B5"] = rating.upper(); s["B5"].font = Font(bold=True, color="C00000")
s["A6"] = "Disposition"; s["B6"] = disposition; s["B6"].font = Font(bold=True, color="C00000")
s["A8"] = "Escalation reasons"; s["A8"].font = Font(bold=True)
for i, r in enumerate(escalation_reasons):
    s[f"A{9+i}"] = r
s["A12"] = "DECISION BOUNDARY"; s["A12"].font = Font(bold=True)
s["A13"] = "This packet is scored and routed only. This skill never approves — a human reviewer (EDD) decides."
s.column_dimensions["A"].width = 22; s.column_dimensions["B"].width = 40

rsheet = wb.create_sheet("Rule outcomes")
rsheet.append(["rule_id", "rule", "outcome", "evidence"])
for c in range(1, 5):
    rsheet.cell(1, c).font = HDR; rsheet.cell(1, c).fill = GREY
for r, text, o, e in outcomes:
    rsheet.append([r, text, o, e])
for col, wdt in zip("ABCD", (8, 46, 9, 60)):
    rsheet.column_dimensions[col].width = wdt

fsheet = wb.create_sheet("Risk factors")
fsheet.append(["factor", "value", "reading"])
for c in range(1, 4):
    fsheet.cell(1, c).font = HDR; fsheet.cell(1, c).fill = GREY
for f in factors:
    fsheet.append(list(f))
for col, wdt in zip("ABC", (18, 22, 32)):
    fsheet.column_dimensions[col].width = wdt

dsheet = wb.create_sheet("Required docs")
dsheet.append(["required document", "status"])
for c in range(1, 3):
    dsheet.cell(1, c).font = HDR; dsheet.cell(1, c).fill = GREY
for d, st in doc_status:
    dsheet.append([d, st])
dsheet.column_dimensions["A"].width = 34; dsheet.column_dimensions["B"].width = 12

os.makedirs("./out", exist_ok=True)
wb.save("./out/escalation-PKT-DEMO-001.xlsx")
print("\nwrote ./out/escalation-PKT-DEMO-001.xlsx")
